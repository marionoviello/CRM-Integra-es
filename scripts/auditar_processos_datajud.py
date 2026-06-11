#!/usr/bin/env python3
"""Audita a carteira de processos do Juridiq contra o DataJud (CNJ).

Pra cada processo ativo no Juridiq, consulta a API pública oficial do
CNJ (DataJud) e compara a última movimentação dos dois lados.

Buckets do relatório:
  - MONITORAMENTO_FALHOU: tribunal tem movimentação MAIS RECENTE que o
    Juridiq (>7 dias de diferença) → religar monitoramento, há
    andamentos perdidos.
  - EM_DIA: Juridiq está igual ou à frente do DataJud.
  - PARADO_HA_MUITO: ambos os lados sem movimentação há >365 dias →
    avaliar arquivamento / prescrição intercorrente.
  - NAO_ENCONTRADO: DataJud não devolve o processo (segredo de justiça,
    número errado, ou tribunal não coberto).

Saída: relatorio_auditoria_processos.xlsx (4 abas) + resumo no stdout.

Uso:
    JURIDIQ_API_KEY=... uv run --with httpx --with openpyxl \
        python scripts/auditar_processos_datajud.py

API DataJud: chave PÚBLICA oficial publicada pelo CNJ em
https://datajud-wiki.cnj.jus.br/api-publica/acesso (gratuita).
Rate limit documentado: 120 req/min — usamos throttle 0.6s.
"""

from __future__ import annotations

import datetime
import os
import re
import sys
import time

import httpx

DATAJUD_APIKEY = "cDZHYzlZa0JadVREZDJCendQbXY6SkJlTzNjLV9TRENyQk1RdnFKZGRQdw=="
JURIDIQ_BASE = "https://api.juridiq.com.br"
THROTTLE_S = 0.6

# Mapa J.TR do número CNJ → alias do endpoint DataJud.
# Formato CNJ: NNNNNNN-DD.AAAA.J.TR.OOOO
_TRIBUNAIS = {
    # Justiça Estadual (J=8)
    ("8", "26"): "tjsp", ("8", "19"): "tjrj", ("8", "13"): "tjmg",
    ("8", "21"): "tjrs", ("8", "16"): "tjpr", ("8", "24"): "tjsc",
    ("8", "07"): "tjdft", ("8", "05"): "tjba", ("8", "17"): "tjpe",
    ("8", "06"): "tjce", ("8", "09"): "tjgo", ("8", "08"): "tjes",
    ("8", "15"): "tjpb", ("8", "20"): "tjrn", ("8", "02"): "tjal",
    ("8", "25"): "tjse", ("8", "27"): "tjto", ("8", "10"): "tjma",
    ("8", "22"): "tjro", ("8", "01"): "tjac", ("8", "04"): "tjam",
    ("8", "14"): "tjpa", ("8", "03"): "tjap", ("8", "23"): "tjrr",
    ("8", "11"): "tjmt", ("8", "12"): "tjms", ("8", "18"): "tjpi",
    # Justiça Federal (J=4)
    ("4", "01"): "trf1", ("4", "02"): "trf2", ("4", "03"): "trf3",
    ("4", "04"): "trf4", ("4", "05"): "trf5", ("4", "06"): "trf6",
    # Justiça do Trabalho (J=5) — trt2 SP capital, trt15 Campinas
    ("5", "02"): "trt2", ("5", "15"): "trt15",
    # Superiores
    ("3", "00"): "stj",
}


def _alias_datajud(process_number: str) -> str | None:
    m = re.match(
        r"^\d{7}-?\d{2}\.?(\d{4})\.?(\d)\.?(\d{2})\.?\d{4}$",
        process_number.strip(),
    )
    if not m:
        return None
    return _TRIBUNAIS.get((m.group(2), m.group(3)))


def _so_digitos(s: str) -> str:
    return re.sub(r"\D", "", s)


def consultar_datajud(
    client: httpx.Client, process_number: str,
) -> tuple[str | None, str]:
    """Retorna (data_ultima_movimentacao_iso | None, status_da_consulta)."""
    alias = _alias_datajud(process_number)
    if not alias:
        return None, "tribunal_nao_mapeado"
    try:
        r = client.post(
            f"https://api-publica.datajud.cnj.jus.br/api_publica_{alias}/_search",
            headers={
                "Authorization": f"APIKey {DATAJUD_APIKEY}",
                "Content-Type": "application/json",
            },
            json={
                "query": {"match": {"numeroProcesso": _so_digitos(process_number)}},
                "size": 5,
            },
        )
        if r.status_code >= 400:
            return None, f"http_{r.status_code}"
        hits = r.json().get("hits", {}).get("hits", [])
        if not hits:
            return None, "nao_encontrado"
        # Processo pode ter registro em múltiplas instâncias — pega a
        # movimentação mais recente entre todos os hits.
        datas = []
        for h in hits:
            for mov in h["_source"].get("movimentos", []) or []:
                dh = mov.get("dataHora")
                if dh:
                    datas.append(dh)
        if not datas:
            return None, "sem_movimentos"
        return max(datas), "ok"
    except httpx.HTTPError as exc:
        return None, f"erro_{type(exc).__name__}"


def main() -> int:
    api_key = os.environ.get("JURIDIQ_API_KEY", "").strip()
    if not api_key:
        print("ERRO: defina JURIDIQ_API_KEY", file=sys.stderr)
        return 2

    jq = httpx.Client(
        base_url=JURIDIQ_BASE,
        headers={"x-juridiq-api-key": api_key},
        timeout=30.0,
    )
    dj = httpx.Client(timeout=30.0)

    suits, page = [], 1
    while True:
        d = jq.get("/lawSuit/", params={"page": page, "limit": 100}).json()
        suits.extend(d["data"])
        if page >= int(d.get("totalPages") or 1):
            break
        page += 1
    jq.close()
    print(f"processos no Juridiq: {len(suits)}")

    agora = datetime.datetime.now(datetime.UTC)
    falhou, em_dia, parado, nao_enc = [], [], [], []

    for i, s in enumerate(suits, 1):
        pn = s.get("processNumber") or ""
        if not pn:
            continue
        dj_data, dj_status = consultar_datajud(dj, pn)
        jq_data = s.get("lastMovementDate")

        jq_dt = None
        if jq_data:
            try:
                jq_dt = datetime.datetime.fromisoformat(
                    str(jq_data).replace("Z", "+00:00")
                )
            except ValueError:
                pass
        dj_dt = None
        if dj_data:
            try:
                dj_dt = datetime.datetime.fromisoformat(
                    str(dj_data).replace("Z", "+00:00")
                )
            except ValueError:
                pass

        linha = {
            "processo": pn,
            "monitoringStatus": s.get("monitoringStatus"),
            "juridiq_ultima_mov": str(jq_data or "")[:10],
            "tribunal_ultima_mov": str(dj_data or "")[:10],
            "consulta": dj_status,
        }

        if dj_dt is None:
            nao_enc.append(linha)
        elif jq_dt is None or (dj_dt - jq_dt).days > 7:
            dias = (dj_dt - jq_dt).days if jq_dt else None
            linha["atraso_dias"] = dias if dias is not None else "sem data no Juridiq"
            falhou.append(linha)
        elif (agora - dj_dt).days > 365:
            linha["parado_dias"] = (agora - dj_dt).days
            parado.append(linha)
        else:
            em_dia.append(linha)

        if i % 25 == 0:
            print(f"  ...{i}/{len(suits)} consultados")
        time.sleep(THROTTLE_S)

    dj.close()

    print()
    print("=== RESULTADO DA AUDITORIA ===")
    print(f"MONITORAMENTO FALHOU (tribunal à frente): {len(falhou)}")
    print(f"EM DIA:                                   {len(em_dia)}")
    print(f"PARADO HÁ >1 ANO (ambos os lados):        {len(parado)}")
    print(f"NÃO ENCONTRADO no DataJud:                {len(nao_enc)}")

    # Planilha de saída
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    abas = [
        ("Monitoramento FALHOU", falhou, "C00000"),
        ("Parado +1 ano", parado, "ED7D31"),
        ("Nao encontrado", nao_enc, "808080"),
        ("Em dia", em_dia, "2E7D32"),
    ]
    wb.remove(wb.active)
    for titulo, linhas, cor in abas:
        ws = wb.create_sheet(titulo)
        if not linhas:
            ws["A1"] = "(vazio)"
            continue
        cols = list(linhas[0].keys())
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color=cor)
            ws.column_dimensions[cell.column_letter].width = 24
        for ri, linha in enumerate(linhas, 2):
            for ci, c in enumerate(cols, 1):
                ws.cell(row=ri, column=ci, value=linha.get(c))
    out = "relatorio_auditoria_processos.xlsx"
    wb.save(out)
    print(f"\nplanilha: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
