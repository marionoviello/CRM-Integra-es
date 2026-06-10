#!/usr/bin/env python3
"""Importa clientes de planilhas .xls (CRM anterior) pro Juridiq.

Uso:
    # 1. Dry-run (default — NÃO cria nada, só mostra o que faria):
    JURIDIQ_API_KEY=... uv run --with xlrd python scripts/importar_clientes_juridiq.py planilha1.xls planilha2.xls

    # 2. Importação de verdade:
    JURIDIQ_API_KEY=... uv run --with xlrd python scripts/importar_clientes_juridiq.py --execute planilha1.xls ...

Proteções:
  - Dedupe interno (mesmo cliente em 2 planilhas) por documento/nome
  - Dedupe contra o Juridiq: baixa TODAS as pessoas existentes e pula
    quem já está lá (por documento normalizado; fallback nome)
  - Throttle 0.3s entre creates
  - Relatório final: criados / pulados / erros (com linha de origem)

Formato esperado (export do CRM anterior, 18 colunas):
  Tipo | Cliente | Cliente Relacionado | Data de cadastro |
  Data de Nascimento/Fundação | Email(s) | Telefone(s) | Documento 1 |
  Documento 2 | Tags | Responsável | UF | Cidade | CEP | Logradouro |
  Número | Complemento | Bairro
"""

from __future__ import annotations

import argparse
import datetime
import os
import re
import sys
import time
import unicodedata

import httpx
import xlrd

BASE_URL = "https://api.juridiq.com.br"
THROTTLE_S = 0.3


# --- Normalização ----------------------------------------------------------

def _norm_doc(raw: object) -> str:
    """'41107268834.0' → '41107268834'; zero-pad CPF(11)/CNPJ(14)."""
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if len(digits) in (9, 10):       # CPF que perdeu zeros à esquerda
        digits = digits.zfill(11)
    elif len(digits) in (12, 13):    # CNPJ idem
        digits = digits.zfill(14)
    return digits


def _norm_phone(raw: str) -> tuple[str, list[str]]:
    """'(27) 3319-8035; (12) 99...' → ('552733198035', [outros...])."""
    parts = [p.strip() for p in str(raw).split(";") if p.strip()]
    norm = []
    for p in parts:
        d = re.sub(r"\D", "", p)
        if not d:
            continue
        if len(d) in (10, 11):       # DDD + número, sem DDI
            d = "55" + d
        norm.append(d)
    if not norm:
        return "", []
    return norm[0], norm[1:]


def _norm_emails(raw: str) -> tuple[str, list[str]]:
    parts = [p.strip().lower() for p in str(raw).split(";") if p.strip()]
    parts = [p for p in parts if "@" in p]
    if not parts:
        return "", []
    return parts[0], parts[1:]


def _norm_name(raw: str) -> str:
    """Lowercase sem acentos pra comparação de dedupe."""
    s = unicodedata.normalize("NFKD", str(raw))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _parse_date(cell_value: object, datemode: int) -> str:
    """Excel float ou 'dd/mm/yyyy' → 'YYYY-MM-DD' (vazio se inválido)."""
    if isinstance(cell_value, float) and cell_value > 0:
        try:
            dt = xlrd.xldate_as_datetime(cell_value, datemode)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return ""
    s = str(cell_value).strip()
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return ""


# --- Leitura das planilhas --------------------------------------------------

def ler_planilhas(paths: list[str]) -> list[dict]:
    """Lê .xls (xlrd, BIFF imperfeito do CRM antigo) e .xlsx (openpyxl,
    incluindo o template padrão template_importacao_juridiq.xlsx)."""
    registros = []
    for path in paths:
        if path.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            sh = wb["Clientes"] if "Clientes" in wb.sheetnames else wb.active
            rows = list(sh.iter_rows(values_only=True))
            header = [str(c or "").strip() for c in rows[0]]
            for i, valores in enumerate(rows[1:], start=2):
                row = dict(zip(header, ("" if v is None else v for v in valores)))
                row["_origem"] = f"{os.path.basename(path)}:{i}"
                row["_datemode"] = 0
                registros.append(row)
        else:
            wb = xlrd.open_workbook(path, ignore_workbook_corruption=True)
            sh = wb.sheets()[0]
            header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
            for r in range(1, sh.nrows):
                row = dict(zip(header, (sh.cell_value(r, c) for c in range(sh.ncols))))
                row["_origem"] = f"{os.path.basename(path)}:{r + 1}"
                row["_datemode"] = wb.datemode
                registros.append(row)
    return registros


# Campos aceitos no formato PADRÃO (template_importacao_juridiq.xlsx —
# headers = nomes técnicos da API).
_CAMPOS_PADRAO = [
    "name", "personType", "personOrigin", "email", "phone", "document",
    "rg", "birthDate", "maritalStatus", "nationality", "profession",
    "zipCode", "state", "city", "neighborhood", "streetAndNumber",
    "addressComplement", "clientDiscoverOffice", "annotation", "code",
    "isPrivate",
]


def _montar_payload_padrao(row: dict) -> dict:
    """Template padrão: headers já são os campos da API — passa direto,
    normalizando phone/document/email."""
    body: dict = {}
    for campo in _CAMPOS_PADRAO:
        v = str(row.get(campo, "") or "").strip()
        if not v or v in ("-", "--"):
            continue
        if campo == "phone":
            v, _ = _norm_phone(v)
        elif campo == "document":
            v = _norm_doc(v)
        elif campo == "email":
            v = v.lower()
        elif campo == "isPrivate":
            body[campo] = v.upper() in ("TRUE", "1", "SIM", "VERDADEIRO")
            continue
        if v:
            body[campo] = v
    body.setdefault("personOrigin", "Cliente")
    body.setdefault("personType", "física")
    return body


def montar_payload(row: dict) -> dict:
    """Linha da planilha → body do POST /person/.

    Detecta o formato: header 'name' = template padrão (campos da API
    direto); senão, formato do CRM antigo (colunas em português).
    """
    if "name" in row:
        return _montar_payload_padrao(row)

    def col(*names: str) -> str:
        """Valor da primeira coluna cujo nome começa com um dos prefixos.

        Valores-placeholder do CRM antigo ('-', '--') contam como vazio.
        """
        for k in row:
            if any(str(k).startswith(n) for n in names):
                v = str(row[k]).strip()
                if v and v not in ("-", "--"):
                    return v
        return ""

    tipo = col("Tipo")
    phone, outros_tels = _norm_phone(col("Telefone"))
    email, outros_emails = _norm_emails(col("Email"))
    doc1 = _norm_doc(col("Documento 1"))
    doc2 = str(row.get("Documento 2", "")).strip()
    if doc2.endswith(".0"):
        doc2 = doc2[:-2]
    nasc = _parse_date(
        next((row[k] for k in row if str(k).startswith("Data de Nascimento")), ""),
        row["_datemode"],
    )
    cadastro = _parse_date(
        next((row[k] for k in row if str(k).startswith("Data de cadastro")), ""),
        row["_datemode"],
    )

    logradouro = col("Logradouro")
    numero = col("Número", "Numero", "Número")
    if numero.endswith(".0"):  # número de rua veio como float do Excel
        numero = numero[:-2]
    street = f"{logradouro}, {numero}" if logradouro and numero else logradouro

    notas = ["Importado do CRM anterior (10/06/2026)."]
    if col("Responsável", "Respons"):
        notas.append(f"Responsável original: {col('Responsável', 'Respons')}")
    if cadastro:
        notas.append(f"Cadastro original: {cadastro}")
    if col("Tags"):
        notas.append(f"Tags: {col('Tags')}")
    if outros_tels:
        notas.append(f"Outros telefones: {', '.join(outros_tels)}")
    if outros_emails:
        notas.append(f"Outros emails: {', '.join(outros_emails)}")

    body: dict = {
        "name": col("Cliente"),
        "personOrigin": "Cliente",
        "personType": "jurídica" if tipo == "PJ" else "física",
        "annotation": "\n".join(notas),
    }
    if phone:
        body["phone"] = phone
    if email:
        body["email"] = email
    if doc1:
        body["document"] = doc1
    if doc2 and tipo == "PF":
        body["rg"] = doc2
    if nasc:
        body["birthDate"] = nasc
    if col("CEP"):
        body["zipCode"] = re.sub(r"\D", "", col("CEP"))
    if col("UF"):
        body["state"] = col("UF")
    if col("Cidade"):
        body["city"] = col("Cidade")
    if col("Bairro"):
        body["neighborhood"] = col("Bairro")
    if street:
        body["streetAndNumber"] = street
    if col("Complemento"):
        body["addressComplement"] = col("Complemento")
    return body


# --- Juridiq ----------------------------------------------------------------

def baixar_existentes(client: httpx.Client) -> tuple[set, set]:
    """Todas as pessoas do Juridiq → (set de documentos, set de nomes norm)."""
    docs, nomes = set(), set()
    page = 1
    while True:
        resp = client.get("/person/", params={"page": page, "limit": 100})
        resp.raise_for_status()
        data = resp.json()
        for p in data.get("data", []):
            if p.get("document"):
                docs.add(_norm_doc(p["document"]))
            if p.get("name"):
                nomes.add(_norm_name(p["name"]))
        if page >= int(data.get("totalPages") or 1):
            break
        page += 1
    return docs, nomes


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("planilhas", nargs="+")
    ap.add_argument("--execute", action="store_true",
                    help="cria de verdade (default: dry-run)")
    ap.add_argument("--skip", action="append", default=[],
                    help="documento OU fragmento de nome a excluir "
                         "(pode repetir)")
    args = ap.parse_args()

    api_key = os.environ.get("JURIDIQ_API_KEY", "").strip()
    if not api_key:
        print("ERRO: defina JURIDIQ_API_KEY", file=sys.stderr)
        return 2

    registros = ler_planilhas(args.planilhas)
    print(f"planilhas: {len(args.planilhas)} | linhas lidas: {len(registros)}")

    # Dedupe interno (entre planilhas): documento > nome.
    vistos: set = set()
    unicos = []
    dups_internos = 0
    for row in registros:
        body = montar_payload(row)
        chave = body.get("document") or _norm_name(body["name"])
        if chave in vistos:
            dups_internos += 1
            continue
        vistos.add(chave)
        unicos.append((row, body))
    print(f"dedupe interno: {dups_internos} duplicados | {len(unicos)} únicos")

    client = httpx.Client(
        base_url=BASE_URL,
        headers={"x-juridiq-api-key": api_key},
        timeout=20.0,
    )

    print("baixando pessoas existentes do Juridiq pra dedupe...")
    docs_exist, nomes_exist = baixar_existentes(client)
    print(f"existentes no Juridiq: {len(docs_exist)} documentos, {len(nomes_exist)} nomes")

    skips = [s.strip().lower() for s in args.skip]

    def _skipped(body: dict) -> bool:
        doc = body.get("document", "")
        nome = _norm_name(body["name"])
        return any(s == doc or s in nome for s in skips)

    a_criar, pulados, excluidos = [], [], []
    for row, body in unicos:
        doc = body.get("document", "")
        if _skipped(body):
            excluidos.append(body["name"])
        elif doc and doc in docs_exist:
            pulados.append((row["_origem"], body["name"], "documento já existe"))
        elif _norm_name(body["name"]) in nomes_exist:
            pulados.append((row["_origem"], body["name"], "nome já existe"))
        else:
            a_criar.append((row, body))

    if excluidos:
        print(f"excluídos por --skip: {excluidos}")

    print()
    print(f"=== RESUMO {'(DRY-RUN — nada será criado)' if not args.execute else '(EXECUTANDO)'} ===")
    print(f"a criar:  {len(a_criar)}")
    print(f"pulados:  {len(pulados)} (já existem no Juridiq)")
    print()
    print("=== amostra do que será criado (5 primeiros) ===")
    for _, body in a_criar[:5]:
        print(f"- {body['name']} | {body.get('personType')} | tel={body.get('phone','—')} "
              f"| email={body.get('email','—')} | doc={body.get('document','—')}")

    if not args.execute:
        print()
        print("Dry-run concluído. Rode com --execute pra importar.")
        client.close()
        return 0

    print()
    criados, erros = 0, []
    for row, body in a_criar:
        try:
            resp = client.post("/person/", json=body)
            if resp.status_code >= 400:
                erros.append((row["_origem"], body["name"],
                              f"HTTP {resp.status_code}: {resp.text[:120]}"))
            else:
                criados += 1
                if criados % 25 == 0:
                    print(f"  ...{criados} criados")
        except Exception as exc:
            erros.append((row["_origem"], body["name"], str(exc)[:120]))
        time.sleep(THROTTLE_S)

    client.close()
    print()
    print("=== RELATÓRIO FINAL ===")
    print(f"criados: {criados}")
    print(f"pulados: {len(pulados)}")
    print(f"erros:   {len(erros)}")
    for origem, nome, err in erros:
        print(f"  ERRO {origem} {nome!r}: {err}")
    return 0 if not erros else 1


if __name__ == "__main__":
    sys.exit(main())
